import openpyxl
import pytz
from openpyxl.styles import Font, PatternFill, Alignment
from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponse, JsonResponse
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.contrib.auth.decorators import login_required
from django.views.generic import TemplateView, DetailView
from apps.fleet.models import Machine, KTGHistory, RepairLog, KTGMonthResult
from apps.fleet.models import Machine, RepairLog
from datetime import timedelta
from django.utils import timezone as dj_timezone
from apps.fleet.models import EngineHoursLog


# Create your views here.
class DashboardView(LoginRequiredMixin, TemplateView):
    template_name = "fleet/dashboard.html"


class MachineDetailView(LoginRequiredMixin, UserPassesTestMixin, DetailView):
    """
    Детальная страница машины.
    DetailView автоматически берёт машину по pk из URL
    и передаёт в шаблон как object или machine.
    """

    model = Machine
    template_name = "fleet/machine_detail.html"
    context_object_name = "machine"  # в шаблоне будет {{ machine }}

    def get_context_data(self, **kwargs):
        """
        Добавляем дополнительные данные в контекст шаблона.
        super() вызывает родительский метод — он добавляет machine.
        Мы добавляем историю ремонтов и результаты КТГ.
        """
        context = super().get_context_data(**kwargs)
        machine = self.object

        # История ремонтов — последние 10
        context["repairs"] = (
            RepairLog.objects.filter(machine=machine)
            .select_related("user")
            .order_by("-created_at")[:10]
        )

        # Результаты КТГ за периоды
        context["ktg_results"] = KTGMonthResult.objects.filter(
            machine=machine
        ).order_by("-period_end")[
            :12
        ]  # последние 12 месяцев

        # История КТГ для графика — последние 100 точек
        context["ktg_history"] = KTGHistory.objects.filter(machine=machine).order_by(
            "-created_at"
        )[:100]

        return context

    def test_func(self):
        return self.request.user.role != "viewer"


def get_machine_timezone(machine):
    """Возвращает timezone участка машины или UTC"""
    if machine.section and machine.section.timezone:
        try:
            return pytz.timezone(machine.section.timezone)
        except pytz.exceptions.UnknownTimeZoneError:
            pass
    return pytz.utc


def localize_dt(dt, tz):
    """Конвертирует datetime в нужный timezone"""
    if dt is None:
        return None
    return dt.astimezone(tz)


@login_required
def export_repairs_excel(request, pk):
    """
    Выгрузка истории ремонтов машины в Excel.
    pk — ID машины.
    """

    machine = Machine.objects.select_related("section").get(pk=pk)
    tz = get_machine_timezone(machine)
    repairs = (
        RepairLog.objects.filter(machine=machine)
        .select_related("user")
        .order_by("-created_at")
    )

    # Создаём книгу
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "История ремонтов"

    # Стили заголовков
    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="1F2937")
    header_align = Alignment(horizontal="center", vertical="center")

    # Заголовки
    headers = [
        "Машина",
        "Бортовой №",
        "Кто отправил",
        "Начало ремонта",
        "Завершён",
        "Продолжительность (часы)",
        "Статус",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Данные
    for row, repair in enumerate(repairs, 2):
        # Начало ремонта
        started = localize_dt(repair.repair_started_at or repair.created_at, tz)
        finished = localize_dt(repair.finished_at, tz)

        # Продолжительность в часах
        if started and finished:
            duration = repair.finished_at - (
                repair.repair_started_at or repair.created_at
            )
            duration_hours = round(duration.total_seconds() / 3600, 2)
        else:
            duration_hours = "—"

        ws.cell(row=row, column=1, value=machine.name)
        ws.cell(row=row, column=2, value=machine.board_number)
        ws.cell(
            row=row, column=3, value=repair.user.get_full_name() if repair.user else "—"
        )
        ws.cell(
            row=row,
            column=4,
            value=started.strftime("%d.%m.%Y %H:%M") if started else "—",
        )
        ws.cell(
            row=row,
            column=5,
            value=finished.strftime("%d.%m.%Y %H:%M") if finished else "—",
        )
        ws.cell(row=row, column=6, value=duration_hours)
        ws.cell(row=row, column=7, value="Активен" if repair.is_active else "Завершён")

    # Ширина колонок
    col_widths = [20, 15, 25, 20, 20, 25, 12]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

    # Отдаём файл
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f'attachment; filename="repairs_{machine.board_number}.xlsx"'
    )
    wb.save(response)
    return response


@login_required
def export_ktg_results_excel(request, pk):
    """
    Выгрузка результатов КТГ по периодам в Excel.
    """
    from apps.fleet.models import Machine, KTGMonthResult

    machine = Machine.objects.select_related("section").get(pk=pk)
    tz = get_machine_timezone(machine)
    results = KTGMonthResult.objects.filter(machine=machine).order_by("-period_end")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Результаты КТГ"

    # Стили заголовков
    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="1F2937")
    header_align = Alignment(horizontal="center", vertical="center")

    headers = [
        "Машина",
        "Бортовой №",
        "Период начало",
        "Период конец",
        "Итоговый КТГ",
        "КТГ %",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    for row, result in enumerate(results, 2):
        period_start = localize_dt(result.period_start, tz)
        period_end = localize_dt(result.period_end, tz)

        ws.cell(row=row, column=1, value=machine.name)
        ws.cell(row=row, column=2, value=machine.board_number)
        ws.cell(
            row=row,
            column=3,
            value=period_start.strftime("%d.%m.%Y") if period_start else "—",
        )
        ws.cell(
            row=row,
            column=4,
            value=period_end.strftime("%d.%m.%Y") if period_end else "—",
        )
        ws.cell(row=row, column=5, value=round(result.value, 6))
        ws.cell(row=row, column=6, value=f"{round(result.value * 100, 4)}%")

    col_widths = [20, 15, 18, 18, 15, 12]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f'attachment; filename="ktg_results_{machine.board_number}.xlsx"'
    )
    wb.save(response)
    return response


@login_required
def api_ktg_history(request, pk):
    """
    API для графика истории КТГ.
    Возвращает JSON с метками времени и значениями КТГ.
    """
    machine = Machine.objects.select_related("section").get(pk=pk)
    tz = get_machine_timezone(machine)

    # Берём последние 200 точек из истории
    history = KTGHistory.objects.filter(machine=machine).order_by("created_at")[:200]

    labels = []
    values = []

    for point in history:
        # Конвертируем время в timezone участка
        local_time = point.created_at.astimezone(tz)
        labels.append(local_time.strftime("%d.%m %H:%M"))
        values.append(round(point.value * 100, 4))  # в процентах

    return JsonResponse(
        {
            "labels": labels,
            "values": values,
        }
    )


@login_required
def api_repairs_history(request, pk):
    """
    API для графика продолжительности ремонтов.
    Параметр period — количество дней (7/30/90/180/365).
    """
    machine = Machine.objects.select_related("section").get(pk=pk)
    tz = get_machine_timezone(machine)

    # Читаем период из GET параметра — по умолчанию 30 дней
    period = int(request.GET.get("period", 30))
    date_from = dj_timezone.now() - timedelta(days=period)

    repairs = RepairLog.objects.filter(
        machine=machine,
        created_at__gte=date_from,
        finished_at__isnull=False,  # только завершённые
    ).order_by("created_at")

    labels = []
    values = []

    for repair in repairs:
        started = repair.repair_started_at or repair.created_at
        local_time = started.astimezone(tz)
        labels.append(local_time.strftime("%d.%m.%Y"))

        # Продолжительность в часах
        duration = repair.finished_at - started
        hours = round(duration.total_seconds() / 3600, 2)
        values.append(hours)

    return JsonResponse(
        {
            "labels": labels,
            "values": values,
            "period": period,
        }
    )


@login_required
def update_engine_hours(request, pk):
    """
    Обновляем моточасы машины.
    Новое значение должно быть больше предыдущего.
    Записываем лог — кто и когда внёс.
    """
    if request.method == 'POST':
        machine = Machine.objects.get(pk=pk)
        new_hours = request.POST.get('engine_hours')

        if new_hours:
            new_hours = float(new_hours)

            # Проверяем — новое значение должно быть больше текущего
            if new_hours <= machine.engine_hours:
                # Передаём ошибку в шаблон через сессию
                request.session['hours_error'] = (
                    f'Новое значение ({new_hours}) должно быть '
                    f'больше текущего ({machine.engine_hours})'
                )
                return redirect('fleet:machine_detail', pk=pk)

            # Сохраняем новое значение
            machine.engine_hours = new_hours
            machine.save()

            # Записываем в лог
            EngineHoursLog.objects.create(
                machine=machine,
                value=new_hours,
                user=request.user
            )

            # Очищаем ошибку если была
            request.session.pop('hours_error', None)

    return redirect('fleet:machine_detail', pk=pk)